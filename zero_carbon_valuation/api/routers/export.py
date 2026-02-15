"""
导出API路由
提供Excel和JSON格式的项目导出功能
"""
from fastapi import APIRouter, HTTPException
from typing import Dict, Any
import io
import json
from datetime import datetime

from models import ExportRequest

router = APIRouter(prefix="/api/v1/export", tags=["Export"])


@router.post("/project/json")
async def export_project_json(req: ExportRequest) -> Dict[str, Any]:
    """
    导出项目为JSON格式

    返回包含元数据和项目数据的JSON对象
    """
    try:
        export_data = {
            "metadata": {
                "version": "1.0",
                "exportedAt": datetime.now().isoformat(),
                "appName": "零碳项目收益评估软件",
                "appVersion": "1.0.0",
                "projectName": req.project_name
            },
            "project": req.project_data
        }

        return export_data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.post("/project/excel")
async def export_project_excel(req: ExportRequest):
    """
    导出项目为Excel格式

    生成包含项目摘要、各模块详细数据、财务分析的Excel文件

    返回: 文件下载响应
    """
    try:
        # 尝试导入openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from fastapi.responses import Response
        except ImportError:
            raise HTTPException(status_code=500, detail="Excel导出需要安装openpyxl库")

        # 创建工作簿
        wb = Workbook()

        # Sheet 1: 项目摘要
        ws_summary = wb.active
        ws_summary.title = "项目摘要"

        # 标题
        ws_summary["A1"] = "零碳项目收益评估报告"
        ws_summary["A1"].font = Font(size=16, bold=True)
        ws_summary["A1"].alignment = Alignment(horizontal="center")

        ws_summary["A2"] = f"项目名称: {req.project_name}"
        ws_summary["A3"] = f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # 项目基本信息
        ws_summary["A5"] = "项目基本信息"
        ws_summary["A5"].font = Font(bold=True)

        base_info = req.project_data.get("projectBaseInfo", {})
        info_data = [
            ("项目名称", base_info.get("name", "")),
            ("项目类型", base_info.get("type", "")),
            ("省份", base_info.get("province", "")),
            ("城市", base_info.get("city", "")),
        ]

        row = 6
        for key, value in info_data:
            ws_summary[f"A{row}"] = key
            ws_summary[f"B{row}"] = value
            row += 1

        # 模块汇总
        ws_summary["A" + str(row + 1)] = "模块汇总"
        ws_summary["A" + str(row + 1)].font = Font(bold=True)

        ws_summary["A" + str(row + 2)] = "模块名称"
        ws_summary["B" + str(row + 2)] = "状态"
        ws_summary["C" + str(row + 2)] = "投资(万元)"
        ws_summary["D" + str(row + 2)] = "年收益(万元)"
        ws_summary["E" + str(row + 2)] = "ROI(%)"

        # 列表头样式
        for col in ["A", "B", "C", "D", "E"]:
            ws_summary[f"{col}{row + 2}"].font = Font(bold=True)
            ws_summary[f"{col}{row + 2}"].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        modules = req.project_data.get("modules", {})
        row += 3

        for module_id, module in modules.items():
            if module.get("isActive"):
                ws_summary[f"A{row}"] = module.get("name", "")
                ws_summary[f"B{row}"] = "启用" if module.get("isActive") else "禁用"
                ws_summary[f"C{row}"] = module.get("investment", 0)
                ws_summary[f"D{row}"] = module.get("yearlySaving", 0)

                # 计算ROI
                inv = module.get("investment", 0)
                sav = module.get("yearlySaving", 0)
                roi = (sav / inv * 100) if inv > 0 else 0
                ws_summary[f"E{row}"] = round(roi, 2)
                row += 1

        # Sheet 2: 财务分析
        ws_finance = wb.create_sheet("财务分析")

        ws_finance["A1"] = "财务分析"
        ws_finance["A1"].font = Font(size=14, bold=True)

        total_investment = sum(m.get("investment", 0) for m in modules.values() if m.get("isActive"))
        total_saving = sum(m.get("yearlySaving", 0) for m in modules.values() if m.get("isActive"))
        roi = (total_saving / total_investment * 100) if total_investment > 0 else 0
        payback = (total_investment / total_saving) if total_saving > 0 else 0

        finance_data = [
            ("总投资(万元)", total_investment),
            ("年总收益(万元)", total_saving),
            ("投资回报率(ROI)", f"{round(roi, 2)}%"),
            ("回收期(年)", round(payback, 2)),
            ("碳减排(吨)", sum(m.get("params", {}).get("carbonReduction", 0) for m in modules.values() if m.get("isActive")))
        ]

        row = 3
        for key, value in finance_data:
            ws_finance[f"A{row}"] = key
            ws_finance[f"B{row}"] = value
            row += 1

        # Sheet 3: 详细参数
        ws_detail = wb.create_sheet("详细参数")

        ws_detail["A1"] = "各模块详细参数"
        ws_detail["A1"].font = Font(size=14, bold=True)

        row = 3
        for module_id, module in modules.items():
            if module.get("isActive"):
                ws_detail[f"A{row}"] = module.get("name", "")
                ws_detail[f"A{row}"].font = Font(bold=True)
                row += 1

                params = module.get("params", {})
                if isinstance(params, dict):
                    for key, value in params.items():
                        ws_detail[f"A{row}"] = key
                        ws_detail[f"B{row}"] = str(value)
                        row += 1
                row += 1

        # 保存到内存
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # 返回文件
        filename = f"{req.project_name}_报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出Excel失败: {str(e)}")
